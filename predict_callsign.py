"""Predict plausible ENACOM amateur-radio call signs for a pending licence.

Consumes the output of ``main.py`` (``output/listado.xlsx``, or the intermediate
``parsed_list.txt``) and ranks the call signs ENACOM could still assign to an
applicant at a given address.

The rules encoded here come from the *Reglamento General de Radioaficionados*
approved by ENACOM Resolución RESOL-2017-3635-APN-ENACOM#MM, Anexo
IF-2017-25743421-APN-DNPYC#ENACOM:

* 8.1 - a call sign is ``prefijo + numero + sufijo``. The prefix/number pair
  comes from the international series of Appendix 42 RR; the *first letter(s)
  of the suffix* encode the geographic division (table reproduced below).
* 8.2 - the geographic suffix follows the fixed station's address (8.2.1), or
  the DNI address when there is no fixed station (8.2.2).
* 8.3 - a two-letter suffix is only granted, on request, to SUPERIOR or
  ESPECIAL category holders. Everyone else gets three letters.
* 4.10 / 8.4 - lapsed and deceased holders' call signs are eventually
  reassigned, so free slots inside otherwise-saturated blocks are real
  candidates, not just the advancing frontier.

The regulation deliberately leaves the actual choice to ENACOM ("El criterio de
selección y asignación ... será facultad y estará bajo la responsabilidad de la
Autoridad de Aplicación", 8.2), so the ranking below is empirical: it models
where ENACOM has *recently* been assigning within the applicant's division and
scores the free slots accordingly.

One empirical caveat drives the shape of the output. Within an active block
(prefix + numeral + first two suffix letters) ENACOM hands out the final letter
in no discernible order - LU3DQ* was filled V, W, I, A, K, B, E, X, Y, D, ...
So the *block* is predictable; the last letter is effectively a lottery among
whatever slots are free. The report therefore ranks blocks first and lists each
block's free call signs as equally likely draws from it.
"""

from __future__ import annotations

import argparse
import json
import re
import string
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date
from itertools import product
from pathlib import Path
from typing import Iterator, Sequence

CALLSIGN_RE = re.compile(r"^([A-Z][A-Z0-9])(\d)([A-Z]{2,3})$")
DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")

DEFAULT_XLSX = Path("output/listado.xlsx")
DEFAULT_TXT = Path("parsed_list.txt")
DEFAULT_REPORT_XLSX = Path("output/senales_disponibles.xlsx")

ALPHABET = string.ascii_uppercase

#: Reglamento 8.1 takes the prefix/number pair from Appendix 42 RR. All nine
#: numerals are in use nationwide and carry no category or geographic meaning,
#: so they are all live targets even where a small division has yet to see one.
NUMERALS = "123456789"

#: Reglamento 8.1 - geographic division -> inclusive suffix ranges, padded to
#: three characters so two- and three-letter suffixes compare uniformly.
SUFFIX_RANGES: dict[str, tuple[tuple[str, str], ...]] = {
    "CIUDAD AUTONOMA DE BUENOS AIRES": (("AAA", "CZZ"),),
    "BUENOS AIRES": (("DAA", "EZZ"),),
    "SANTA FE": (("FAA", "FZZ"),),
    "CHACO": (("GAA", "GOZ"),),
    "FORMOSA": (("GPA", "GZZ"),),
    "CORDOBA": (("HAA", "HZZ"),),
    "MISIONES": (("IAA", "IZZ"),),
    "ENTRE RIOS": (("JAA", "JZZ"),),
    "TUCUMAN": (("KAA", "KZZ"),),
    "CORRIENTES": (("LAA", "LZZ"),),
    "MENDOZA": (("MAA", "MZZ"),),
    "SANTIAGO DEL ESTERO": (("NAA", "NZZ"),),
    "SALTA": (("OAA", "OZZ"),),
    "SAN JUAN": (("PAA", "PZZ"),),
    "SAN LUIS": (("QAA", "QZZ"),),
    "CATAMARCA": (("RAA", "RZZ"),),
    "LA RIOJA": (("SAA", "SZZ"),),
    "JUJUY": (("TAA", "TZZ"),),
    "LA PAMPA": (("UAA", "UZZ"),),
    "RIO NEGRO": (("VAA", "VZZ"),),
    "CHUBUT": (("WAA", "WZZ"),),
    "SANTA CRUZ": (("XAA", "XOZ"),),
    "TIERRA DEL FUEGO": (("XPA", "XZZ"),),
    "NEUQUEN": (("YAA", "YZZ"),),
    "ANTARTIDA ARGENTINA E ISLAS DEL ATLANTICO SUR": (("ZAA", "ZZZ"),),
}

#: Registry spellings and common shorthand -> canonical division name above.
DIVISION_ALIASES: dict[str, str] = {
    "CIUDAD DE BUENOS AIRES": "CIUDAD AUTONOMA DE BUENOS AIRES",
    "CIUDAD AUTONOMA BUENOS AIRES": "CIUDAD AUTONOMA DE BUENOS AIRES",
    "CABA": "CIUDAD AUTONOMA DE BUENOS AIRES",
    "PROVINCIA DE BUENOS AIRES": "BUENOS AIRES",
    "TIERRA DEL FUEGO A. E I.A.S.": "TIERRA DEL FUEGO",
    "TIERRA DEL FUEGO ANTARTIDA E ISLAS DEL ATLANTICO SUR": "TIERRA DEL FUEGO",
    "ANTARTIDA": "ANTARTIDA ARGENTINA E ISLAS DEL ATLANTICO SUR",
}

#: Reglamento 8.3 - only SUPERIOR/ESPECIAL may request a two-letter suffix.
TWO_LETTER_CATEGORIES = frozenset({"SUPERIOR", "ESPECIAL"})

#: Everyone else, including a first NOVICIO licence, gets three letters.
DEFAULT_SUFFIX_LENGTH = 3

# Score blend across the hierarchy: sub-block (prefix+numeral+2 letters),
# block (prefix+numeral+1 letter) and series (prefix+first letter).
WEIGHT_SUBBLOCK = 0.55
WEIGHT_BLOCK = 0.30
WEIGHT_SERIES = 0.15

#: Normalised recent-activity level at which a block counts as actively being
#: filled rather than dormant.
ACTIVE_ACTIVITY = 0.25

#: Recency weight of a block's newest licence above which the block counts as
#: live regardless of volume. 0.5 == its newest entry is within one half-life.
FRESH_ENOUGH = 0.5

#: A block this full whose recent activity is below ACTIVE_ACTIVITY is dormant,
#: so its free slots are most likely lapsed call signs awaiting reassignment
#: (Reglamento 4.10) rather than virgin space.
STALE_SATURATION = 0.80


@dataclass(frozen=True)
class Licence:
    """One row of the ENACOM public register."""

    holder: str
    category: str
    callsign: str
    vigencia: date | None
    city: str
    province: str
    prefix: str
    numeral: str
    suffix: str

    @property
    def division(self) -> str:
        return canonical_division(self.province)


@dataclass(frozen=True)
class Block:
    """An assignment block: a call sign with its final suffix letter free.

    ENACOM's choice of that last letter looks random, so a block is the finest
    unit at which prediction is meaningful.
    """

    prefix: str
    numeral: str
    letters: str
    score: float
    occupied: int
    free: tuple[str, ...]
    newest: date | None
    kind: str  # "active" | "stale" | "quiet"

    @property
    def label(self) -> str:
        return f"{self.prefix}{self.numeral}{self.letters}"

    @property
    def saturation(self) -> float:
        return self.occupied / len(ALPHABET)


def strip_accents(text: str) -> str:
    decomposed = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")


def normalise(text: str) -> str:
    """Uppercase, de-accent and collapse whitespace for robust matching."""
    return " ".join(strip_accents(text).upper().split())


def canonical_division(raw: str) -> str:
    """Map a registry province string onto a Reglamento 8.1 division name."""
    key = normalise(raw)
    if key in DIVISION_ALIASES:
        return DIVISION_ALIASES[key]
    if key in SUFFIX_RANGES:
        return key
    without_prefix = key.removeprefix("PROVINCIA DE ").strip()
    if without_prefix in SUFFIX_RANGES:
        return without_prefix
    return key


def parse_date(raw: str) -> date | None:
    match = DATE_RE.match(raw.strip())
    if not match:
        return None
    try:
        return date(int(match[1]), int(match[2]), int(match[3]))
    except ValueError:
        return None


def build_licence(fields: Sequence[str]) -> Licence | None:
    """Build a Licence from six ordered registry fields, or None if unusable."""
    holder, category, callsign, vigencia, city, province = (f.strip() for f in fields[:6])
    match = CALLSIGN_RE.match(callsign.upper())
    if not match:
        return None
    return Licence(
        holder=holder,
        category=category.upper(),
        callsign=callsign.upper(),
        vigencia=parse_date(vigencia),
        city=city,
        province=province,
        prefix=match[1],
        numeral=match[2],
        suffix=match[3],
    )


def _rows_from_xlsx(path: Path) -> Iterator[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is in requirements
        raise SystemExit(f"openpyxl is required to read {path}: {exc}") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise SystemExit(f"{path} has no active worksheet.")
    for row in sheet.iter_rows(values_only=True):
        yield ["" if cell is None else str(cell) for cell in row]
    workbook.close()


def _align_row(cells: Sequence[str]) -> list[str] | None:
    """Locate the call-sign column and return the six fields around it.

    ``main.py`` writes rows by splitting text that begins with a newline, so the
    data lands one column to the right of its header. Rather than hard-coding
    that off-by-one, find the column that actually holds a call sign.
    """
    for index, cell in enumerate(cells):
        if CALLSIGN_RE.match(cell.strip().upper()) and index >= 2:
            window = cells[index - 2 : index + 4]
            if len(window) == 6:
                return list(window)
    return None


def load_from_xlsx(path: Path) -> tuple[Licence, ...]:
    licences = (
        build_licence(aligned)
        for aligned in (_align_row(row) for row in _rows_from_xlsx(path))
        if aligned is not None
    )
    return tuple(lic for lic in licences if lic is not None)


def load_from_txt(path: Path) -> tuple[Licence, ...]:
    blocks = path.read_text(encoding="utf-8", errors="replace").split("\n\n")
    licences = []
    for block in blocks:
        fields = [line for line in block.split("\n") if line.strip()]
        if len(fields) < 6:
            continue
        licence = build_licence(fields)
        if licence is not None:
            licences.append(licence)
    return tuple(licences)


def load_register(explicit: Path | None) -> tuple[Licence, ...]:
    """Load the register from an explicit path or the main.py defaults."""
    candidates = [explicit] if explicit else [DEFAULT_XLSX, DEFAULT_TXT]
    for path in candidates:
        if path is None or not path.exists():
            continue
        loader = load_from_xlsx if path.suffix.lower() in {".xlsx", ".xlsm"} else load_from_txt
        licences = loader(path)
        if licences:
            print(f"Loaded {len(licences):,} licences from {path}", file=sys.stderr)
            return licences
        raise SystemExit(f"{path} contained no parseable call signs.")
    raise SystemExit(
        "No register found. Run main.py first, or pass --input "
        f"(looked for {DEFAULT_XLSX} and {DEFAULT_TXT})."
    )


def suffix_sort_key(suffix: str) -> str:
    """Pad a suffix to three characters so ranges compare uniformly."""
    return (suffix + "AAA")[:3]


def suffix_in_division(suffix: str, division: str) -> bool:
    ranges = SUFFIX_RANGES.get(division)
    if not ranges:
        return False
    key = suffix_sort_key(suffix)
    return any(low <= key <= high for low, high in ranges)


def division_suffixes(division: str, length: int) -> tuple[str, ...]:
    """Every suffix of ``length`` letters belonging to ``division``."""
    return tuple(
        "".join(letters)
        for letters in product(ALPHABET, repeat=length)
        if suffix_in_division("".join(letters), division)
    )


def recency_weight(when: date | None, latest: date, half_life_days: float) -> float:
    """Exponential decay so recent assignments dominate the activity model."""
    if when is None or half_life_days <= 0:
        return 0.0
    age_days = (latest - when).days
    if age_days < 0:
        return 1.0
    return 0.5 ** (age_days / half_life_days)


def suffix_length_for(override: int | None) -> int:
    """Reglamento 8.3: three letters unless a two-letter suffix is requested."""
    return override if override is not None else DEFAULT_SUFFIX_LENGTH


def normalised_counter(counter: Counter[tuple[str, ...]]) -> dict[tuple[str, ...], float]:
    if not counter:
        return {}
    peak = max(counter.values())
    if peak <= 0:
        return {}
    return {key: value / peak for key, value in counter.items()}


def classify(normalised_activity: float, saturation: float, freshness: float) -> str:
    """Label a block from its recent volume and how recent its newest entry is.

    Volume alone is misleading: a block holding the single most recent
    assignment in the whole division can still look quiet next to a block that
    absorbed twenty licences. Freshness catches those.
    """
    if freshness >= FRESH_ENOUGH or normalised_activity >= ACTIVE_ACTIVITY:
        return "active"
    if saturation >= STALE_SATURATION:
        return "stale"
    return "quiet"


def rank_blocks(
    licences: Sequence[Licence],
    division: str,
    length: int,
    half_life_days: float,
    prefixes: Sequence[str] | None,
    numerals: Sequence[str] | None,
) -> tuple[Block, ...]:
    """Rank assignment blocks in ``division`` by how recently ENACOM used them."""
    if length < 2:
        raise SystemExit("Suffix length must be at least 2 to form a block.")

    taken = frozenset(lic.callsign for lic in licences)
    local = [lic for lic in licences if lic.division == division and len(lic.suffix) == length]
    if not local:
        raise SystemExit(f"No {length}-letter licences on record for division {division!r}.")

    dated = [lic.vigencia for lic in local if lic.vigencia is not None]
    latest = max(dated) if dated else date.today()

    activity_sub: Counter[tuple[str, ...]] = Counter()
    activity_block: Counter[tuple[str, ...]] = Counter()
    activity_series: Counter[tuple[str, ...]] = Counter()
    occupied: Counter[tuple[str, ...]] = Counter()
    newest: dict[tuple[str, ...], date] = {}

    stem_len = length - 1
    for lic in local:
        sub = (lic.prefix, lic.numeral, lic.suffix[:stem_len])
        occupied[sub] += 1
        weight = recency_weight(lic.vigencia, latest, half_life_days)
        activity_sub[sub] += weight
        activity_block[(lic.prefix, lic.numeral, lic.suffix[0])] += weight
        activity_series[(lic.prefix, lic.suffix[0])] += weight
        if lic.vigencia is not None:
            known = newest.get(sub)
            if known is None or lic.vigencia > known:
                newest[sub] = lic.vigencia

    norm_sub = normalised_counter(activity_sub)
    norm_block = normalised_counter(activity_block)
    norm_series = normalised_counter(activity_series)

    allowed_prefixes = tuple(prefixes) if prefixes else tuple(sorted({lic.prefix for lic in local}))
    allowed_numerals = tuple(numerals) if numerals else tuple(NUMERALS)
    # A block fixes all but the final suffix letter, which is what varies.
    stems = sorted({s[:stem_len] for s in division_suffixes(division, length)})
    tails = list(ALPHABET)

    blocks = []
    for prefix, numeral, stem in product(allowed_prefixes, allowed_numerals, stems):
        sub = (prefix, numeral, stem)
        activity = norm_sub.get(sub, 0.0)
        score = (
            WEIGHT_SUBBLOCK * activity
            + WEIGHT_BLOCK * norm_block.get((prefix, numeral, stem[0]), 0.0)
            + WEIGHT_SERIES * norm_series.get((prefix, stem[0]), 0.0)
        )
        if score <= 0.0:
            continue
        free = tuple(
            f"{prefix}{numeral}{stem}{tail}"
            for tail in tails
            if suffix_in_division(f"{stem}{tail}", division)
            and f"{prefix}{numeral}{stem}{tail}" not in taken
        )
        if not free:
            continue
        count = occupied.get(sub, 0)
        block_newest = newest.get(sub)
        blocks.append(
            Block(
                prefix=prefix,
                numeral=numeral,
                letters=stem,
                score=score,
                occupied=count,
                free=free,
                newest=block_newest,
                kind=classify(
                    activity,
                    count / len(ALPHABET),
                    recency_weight(block_newest, latest, half_life_days),
                ),
            )
        )

    return tuple(sorted(blocks, key=lambda b: (-b.score, b.label)))


def division_report(licences: Sequence[Licence], division: str, length: int) -> dict[str, object]:
    local = [lic for lic in licences if lic.division == division]
    sized = [lic for lic in local if len(lic.suffix) == length]
    space = len(division_suffixes(division, length))
    prefixes = sorted({lic.prefix for lic in sized})
    numerals = sorted({lic.numeral for lic in sized})
    total_space = space * len(prefixes) * len(NUMERALS)
    return {
        "division": division,
        "licences_in_division": len(local),
        "with_requested_suffix_length": len(sized),
        "prefixes_in_use": prefixes,
        "numerals_in_use": numerals,
        "address_space": total_space,
        "free_slots": total_space - len(sized),
        "letters": sorted({lic.suffix[0] for lic in sized}),
    }


def neighbours_in_city(licences: Sequence[Licence], city: str, limit: int) -> tuple[Licence, ...]:
    needle = normalise(city)
    if not needle:
        return ()
    matches = [lic for lic in licences if needle in normalise(lic.city)]
    return tuple(sorted(matches, key=lambda lic: lic.callsign)[:limit])


def wrap_calls(calls: Sequence[str], indent: str, width: int = 72) -> Iterator[str]:
    line = indent
    for call in calls:
        if len(line) + len(call) + 1 > width and line.strip():
            yield line.rstrip()
            line = indent
        line += call + " "
    if line.strip():
        yield line.rstrip()


def print_report(
    report: dict[str, object],
    blocks: Sequence[Block],
    neighbours: Sequence[Licence],
    city: str,
    category: str,
    length: int,
    top: int,
) -> None:
    print()
    print("=" * 72)
    print("PLAUSIBLE ENACOM CALL SIGNS")
    print("=" * 72)
    print(f"  Address division   : {report['division']}  (Reglamento 8.1 / 8.2.1)")
    print(f"  Suffix letters     : {', '.join(str(x) for x in report['letters'])}")
    print(f"  Category assumed   : {category}  ->  {length}-letter suffix (8.3)")
    print(f"  Prefixes in use    : {', '.join(str(x) for x in report['prefixes_in_use'])}")
    print(f"  Licences on record : {report['with_requested_suffix_length']:,} of "
          f"{report['address_space']:,} possible  "
          f"({report['free_slots']:,} slots free)")

    if neighbours:
        print()
        print(f"  Existing licences in {city.upper()}:")
        for lic in neighbours:
            print(f"    {lic.callsign:8} {lic.category:9} {lic.holder}")

    active = [b for b in blocks if b.kind == "active"][:top]
    stale = [b for b in blocks if b.kind == "stale"][:max(3, top // 3)]
    # A quiet division (or a rare suffix length) may have no block that clears
    # the activity bar; fall back to the best-scoring blocks so the report is
    # never silently empty.
    heading = "blocks ENACOM is filling right now"
    if not active:
        active = list(blocks[:top])
        heading = "best-scoring blocks (none currently active)"

    print()
    print(f"  MOST LIKELY - {heading} (top {len(active)}).")
    print("  Within a block the final letter looks randomly chosen, so treat each")
    print("  block's free call signs as roughly equally likely.")
    for block in active:
        newest = block.newest.isoformat() if block.newest else "n/a"
        print()
        print(f"    {block.label}*  score {block.score:.3f}  "
              f"{block.occupied}/26 taken  newest {newest}")
        for line in wrap_calls(block.free, indent="      "):
            print(line)

    if stale:
        print()
        print(f"  LESS LIKELY - dormant but nearly full blocks (top {len(stale)}).")
        print("  Free slots here are probably lapsed licences that ENACOM may")
        print("  reassign (Reglamento 4.10) or reserved ones (8.4).")
        for block in stale:
            newest = block.newest.isoformat() if block.newest else "n/a"
            print()
            print(f"    {block.label}*  score {block.score:.3f}  "
                  f"{block.occupied}/26 taken  newest {newest}")
            for line in wrap_calls(block.free, indent="      "):
                print(line)
    print()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rank plausible ENACOM amateur call signs for a pending licence, "
            "using the register produced by main.py."
        ),
    )
    parser.add_argument("--input", type=Path, default=None,
                        help="Register file (default: output/listado.xlsx, then parsed_list.txt).")
    parser.add_argument("--province", default="BUENOS AIRES",
                        help="Province of the fixed station's address (Reglamento 8.2.1).")
    parser.add_argument("--city", default="",
                        help="City, used only to list existing local licences.")
    parser.add_argument("--category", default="NOVICIO",
                        help="Licence category being issued (default: NOVICIO).")
    parser.add_argument("--suffix-length", type=int, choices=(2, 3), default=None,
                        help="Override the suffix length implied by the category.")
    parser.add_argument("--prefix", action="append", default=None,
                        help="Restrict to a prefix (repeatable), e.g. --prefix LU.")
    parser.add_argument("--numeral", action="append", default=None,
                        help="Restrict to a numeral (repeatable), e.g. --numeral 3.")
    parser.add_argument("--half-life", type=float, default=365.0,
                        help="Recency half-life in days for the activity model (default: 365).")
    parser.add_argument("--top", type=int, default=20, help="How many call signs to show.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of a report.")
    parser.add_argument("--excel", nargs="?", type=Path, const=DEFAULT_REPORT_XLSX, default=None,
                        metavar="PATH",
                        help=f"Also write the full ranking to an .xlsx workbook "
                             f"(default: {DEFAULT_REPORT_XLSX}).")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    licences = load_register(args.input)

    division = canonical_division(args.province)
    if division not in SUFFIX_RANGES:
        raise SystemExit(
            f"Unknown division {args.province!r}. Known: {', '.join(sorted(SUFFIX_RANGES))}"
        )

    category = normalise(args.category)
    length = suffix_length_for(args.suffix_length)
    if length == 2 and category not in TWO_LETTER_CATEGORIES:
        print(
            f"Warning: Reglamento 8.3 reserves two-letter suffixes for "
            f"{'/'.join(sorted(TWO_LETTER_CATEGORIES))}; {category} would not qualify.",
            file=sys.stderr,
        )

    report = division_report(licences, division, length)
    blocks = rank_blocks(
        licences,
        division=division,
        length=length,
        half_life_days=args.half_life,
        prefixes=args.prefix,
        numerals=args.numeral,
    )
    neighbours = neighbours_in_city(licences, args.city, limit=15)

    if args.excel is not None:
        from excel_report import write_report

        path, rows = write_report(args.excel, report, blocks, neighbours, category, length)
        print(f"Wrote {rows:,} candidate call signs across {len(blocks):,} blocks to {path}",
              file=sys.stderr)

    if args.json:
        payload = {
            "summary": report,
            "category": category,
            "suffix_length": length,
            "blocks": [
                {
                    "block": b.label,
                    "score": round(b.score, 4),
                    "kind": b.kind,
                    "occupied": b.occupied,
                    "newest": b.newest.isoformat() if b.newest else None,
                    "free": list(b.free),
                }
                for b in blocks[: args.top]
            ],
            "local_licences": [
                {"callsign": lic.callsign, "category": lic.category, "city": lic.city}
                for lic in neighbours
            ],
        }
        json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
        print()
        return 0

    print_report(report, blocks, neighbours, args.city, category, length, args.top)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
