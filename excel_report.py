"""Write a call-sign prediction to an .xlsx workbook.

Kept apart from ``predict_callsign.py`` so the prediction logic stays free of
presentation concerns. Headers are in Spanish to match ``output/listado.xlsx``,
the workbook ``main.py`` produces.

The terminal report is a summary and honours ``--top``; this workbook is the
full ranked dataset, meant to be sorted and filtered in Excel.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")

#: Block classification -> label shown in the workbook.
KIND_LABELS = {
    "active": "Activo",
    "stale": "Latente",
    "quiet": "Inactivo",
}

BLOCK_HEADERS = (
    "Bloque",
    "Prefijo",
    "Numeral",
    "Letras",
    "Puntaje",
    "Estado",
    "Ocupadas",
    "Libres",
    "Última alta",
    "Señales libres",
)

CALL_HEADERS = (
    "Señal distintiva",
    "Prefijo",
    "Numeral",
    "Sufijo",
    "Bloque",
    "Puntaje",
    "Estado",
)

LOCAL_HEADERS = ("Señal distintiva", "Categoría", "Ciudad", "Titular")


def _style_sheet(sheet: Worksheet, headers: Sequence[str], widths: Sequence[int]) -> None:
    """Apply the shared header styling, freeze panes and autofilter."""
    for column, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_summary(sheet: Worksheet, report: dict[str, object], category: str, length: int) -> None:
    sheet.title = "Resumen"
    rows = (
        ("División (Reglamento 8.1)", report["division"]),
        ("Letras del sufijo", ", ".join(str(x) for x in report["letters"])),
        ("Categoría", category),
        ("Largo del sufijo", f"{length} letras"),
        ("Prefijos en uso", ", ".join(str(x) for x in report["prefixes_in_use"])),
        ("Licencias registradas", report["with_requested_suffix_length"]),
        ("Espacio total de señales", report["address_space"]),
        ("Señales libres", report["free_slots"]),
        ("Generado", date.today().isoformat()),
    )
    _style_sheet(sheet, ("Campo", "Valor"), (32, 52))
    for value_row, (field, value) in enumerate(rows, start=2):
        sheet.cell(row=value_row, column=1, value=field).font = Font(bold=True)
        sheet.cell(row=value_row, column=2, value=value)


def _write_blocks(sheet: Worksheet, blocks: Sequence[object]) -> None:
    _style_sheet(sheet, BLOCK_HEADERS, (10, 8, 9, 8, 9, 10, 10, 8, 13, 60))
    for row, block in enumerate(blocks, start=2):
        sheet.cell(row=row, column=1, value=f"{block.label}*")
        sheet.cell(row=row, column=2, value=block.prefix)
        sheet.cell(row=row, column=3, value=block.numeral)
        sheet.cell(row=row, column=4, value=block.letters)
        sheet.cell(row=row, column=5, value=round(block.score, 4))
        sheet.cell(row=row, column=6, value=KIND_LABELS.get(block.kind, block.kind))
        sheet.cell(row=row, column=7, value=f"{block.occupied}/26")
        sheet.cell(row=row, column=8, value=len(block.free))
        sheet.cell(row=row, column=9, value=block.newest.isoformat() if block.newest else "")
        sheet.cell(row=row, column=10, value=" ".join(block.free))


def _write_callsigns(sheet: Worksheet, blocks: Sequence[object]) -> int:
    _style_sheet(sheet, CALL_HEADERS, (18, 9, 9, 9, 10, 9, 10))
    row = 2
    for block in blocks:
        label = KIND_LABELS.get(block.kind, block.kind)
        score = round(block.score, 4)
        for callsign in block.free:
            sheet.cell(row=row, column=1, value=callsign)
            sheet.cell(row=row, column=2, value=block.prefix)
            sheet.cell(row=row, column=3, value=block.numeral)
            sheet.cell(row=row, column=4, value=callsign[len(block.prefix) + 1:])
            sheet.cell(row=row, column=5, value=f"{block.label}*")
            sheet.cell(row=row, column=6, value=score)
            sheet.cell(row=row, column=7, value=label)
            row += 1
    return row - 2


def _write_local(sheet: Worksheet, neighbours: Sequence[object]) -> None:
    _style_sheet(sheet, LOCAL_HEADERS, (18, 12, 28, 38))
    for row, licence in enumerate(neighbours, start=2):
        sheet.cell(row=row, column=1, value=licence.callsign)
        sheet.cell(row=row, column=2, value=licence.category)
        sheet.cell(row=row, column=3, value=licence.city)
        sheet.cell(row=row, column=4, value=licence.holder)


def write_report(
    path: Path,
    report: dict[str, object],
    blocks: Sequence[object],
    neighbours: Sequence[object],
    category: str,
    length: int,
) -> tuple[Path, int]:
    """Write the workbook and return its path plus the candidate row count."""
    workbook = Workbook()
    summary = workbook.active
    if summary is None:  # pragma: no cover - openpyxl always creates one
        summary = workbook.create_sheet()
    _write_summary(summary, report, category, length)
    _write_blocks(workbook.create_sheet("Bloques"), blocks)
    total = _write_callsigns(workbook.create_sheet("Señales libres"), blocks)
    if neighbours:
        _write_local(workbook.create_sheet("Licencias locales"), neighbours)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path, total
